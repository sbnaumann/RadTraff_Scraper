import sys
import pathlib
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import argparse
import xlrd
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

def search(filename, urls, month, year):
	filename = filename[0]
	urls = urls[0]
	month = month[0]
	year = year[0]

	# Get working directory
	currentDir = str(pathlib.Path().absolute())

	# Setting up the headless chromedriver through options
	options = Options()
	#options.add_argument('--headless')
	options.add_argument('--disable-gpu')
	options.add_argument('--window-size=1920,1200')
	options.add_argument('--ignore-certificate-errors')
	options.add_argument('--disable-extensions')
	options.add_argument('--no-sandbox')
	options.add_argument('--disable-dev-shm-usage')

	# Open a chrome browser
	browser = webdriver.Chrome(currentDir + '/chromedriver', options=options)

	print(urls)

	urlArray = urls.split(",")

	# Open file provided and get search String
	nti_tools_book = openpyxl.load_workbook(filename)
	readSheet = nti_tools_book.active

	print("Loaded a Workbook")

	wb = Workbook()
	workingSheet = wb.active
	sheetHeaders = ["Page Title", "URL"]

	for row in range(1, readSheet.max_row + 1):
		browser.get("http://www.google.com")
		print("Opened a browser")


		print(readSheet.max_row)
		currentTool = readSheet.cell(row, 1).value
		initial_search = currentTool
		print(initial_search)
		if urls:
			print("There were urls")
			for url in urlArray:
				print("adding " + url + " to exception list")
				initial_search = "-" + url + ": " + initial_search
		# prep the string for searching
		initial_search = initial_search.replace(' ','+')

		# Go to google and enter the input into the search bar
		searchField = browser.find_element_by_name("q")
		searchField.send_keys(initial_search)
		searchField.submit()

		now = datetime.datetime.now()
		lastYear = now.year - 1

		if not month and year:
			#  Clicks the Tools button, activates sort dropdowns
			time.sleep(1)
			browser.find_element_by_id('hdtb-tls').click()

			# Need to sort results by last 24, week, month, etc.
			time.sleep(1)
			browser.find_element_by_class_name('hdtb-mn-hd').click()
			time.sleep(1)
			browser.find_element_by_link_text('Past year').click()

		elif not month or year:
			if not month:
				#  Clicks the Tools button, activates sort dropdowns
				time.sleep(1)
				browser.find_element_by_id('hdtb-tls').click()

				time.sleep(1)
				browser.find_element_by_link_text('Custom range...')
				time.sleep(1)
				fromField = browser.find_element_by_id('OouJcb')
				fromField.send_keys(year)
				time.sleep(1)
				browser.find_element_by_name('Go').click()
			if not year:
				#  Clicks the Tools button, activates sort dropdowns
				time.sleep(1)
				browser.find_element_by_id('hdtb-tls').click()

				time.sleep(1)
				browser.find_element_by_link_text('Custom range...')
				time.sleep(1)
				fromField = browser.find_element_by_id('OouJcb')
				fromField.send_keys(month + '/' + lastYear)
				time.sleep(1)
				browser.find_element_by_name('Go').click()

		elif month and year:
				time.sleep(1)
				browser.find_element_by_id('hdtb-tls').click()

				time.sleep(1)
				browser.find_element_by_link_text('Custom range...')
				time.sleep(1)
				fromField = browser.find_element_by_id('OouJcb')
				fromField.send_keys(month + '/' + year)
				time.sleep(1)
				browser.find_element_by_name('Go').click()


		# Search the resulting webpage for URLs
		results = browser.find_elements_by_css_selector("div.g")

		# Print Results
		for result in results:
			link = result.find_element_by_tag_name("a")
			print(link.get_attribute("href"))


		# Search through the returned pages and print them to a file
		if workingSheet.title == "Sheet":
			workingSheet.title = "Sheet " + str(row)
		else:
			workingSheet = wb.create_sheet("Sheet " + str(row))

		workingSheet.cell(1,1,value=readSheet.cell(row,1).value)

		for column in range(0, len(sheetHeaders)):
			workingSheet.cell(2, column+1, value=sheetHeaders[column])

		for i in range(0, workingSheet.max_column):
			workingSheet.column_dimensions[get_column_letter(i+1)].width = 25

		rowCount = 2
		pageLinks = []
		for result in results:
			hyperlinkElement = result.find_element_by_tag_name("a")
			link = hyperlinkElement.get_attribute("href")
			pageLinks.append(link)

		for page in pageLinks:
			browser.get(page)
			try:
				pageTitle = browser.title
			except:
				pageTitle = "Unknown"

			src = browser.page_source
			matches = [currentTool, "nti.org", "NTI.org", "Nuclear Threat Initiative", "nuclear threat initiative"]
			for match in matches:
				if src.find(match) != -1:
					rowCount += 1
					workingSheet.cell(rowCount, 1, value=str(pageTitle))
					workingSheet.cell(rowCount, 2, value=str(page))
					break

	wb.save('results/results.xlsx')
	browser.quit()

def argChecker():
	parser = argparse.ArgumentParser()
	parser.add_argument('filename', type=str, nargs=1)
	parser.add_argument('urls', type=str, nargs=1)
	parser.add_argument('month', type=str, nargs=1)
	parser.add_argument('year', type=str, nargs=1)
	args = parser.parse_args()

	return args

if __name__ == '__main__':
	search(**vars(argChecker()))
